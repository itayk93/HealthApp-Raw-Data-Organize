# Code by Itay Karkason
import pandas as pd
import glob
from datetime import date
from openpyxl import load_workbook
import os
from datetime import datetime
import datetime

today = pd.Timestamp(date.today())

# Define the path of the main folder for use
wb = load_workbook("main_folder_location.xlsx", data_only=True)
sh = wb["Sheet1"]
folder_location = sh["A2"].value
folder_location_output = folder_location + "Output/"
folder_location_script_steps = folder_location + "Script Steps/"

# Define the name of the device holder to show in the data
wb = load_workbook(folder_location + "device_holder.xlsx", data_only=True)
sh = wb["Sheet1"]
holder = sh["A2"].value
kind_of_phone = sh["B2"].value
device_holder = holder + "'s " + kind_of_phone

# Operate Step 1 - Read the input and make raw data files for each type of workout
print("Operate Step 1 - Read the input and make raw data files for each type of workout")
exec(open(folder_location_script_steps + "Health App Script - First Step.py").read())

# Operate Step 2 - Prepare each data file from the raw data files
print("Operate Step 2 - Prepare each data file from the raw data files")
exec(open(folder_location_script_steps + "Health App Script - Second Step.py").read())

# Operate Step 3 - Making Sub-files to Active Energy Burned - By each kind of app
print("Operate Step 3 - Making Sub-files to Active Energy Burned - By each kind of app")
exec(open(folder_location_script_steps + "Health App Script - Third Step.py").read())

# Delete all the work files
len_folder_location = len(folder_location)
all_files = glob.glob(folder_location + "HK*.csv")
for filename in all_files:
    print("'" + filename[len_folder_location:] + "' was deleted")
    os.remove(filename)
